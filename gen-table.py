#!/usr/bin/env python3

import collections
import datetime

import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import quote_sheetname
from openpyxl.utils.cell import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


PAIRINGS_10 = [
    [(1, 10), (2, 9), (3, 8), (4, 7), (5, 6)],
    [(10, 6), (7, 5), (8, 4), (9, 3), (1, 2)],
    [(2, 10), (3, 1), (4, 9), (5, 8), (6, 7)],
    [(10, 7), (8, 6), (9, 5), (1, 4), (2, 3)],
    [(3, 10), (4, 2), (5, 1), (6, 9), (7, 8)],
    [(10, 8), (9, 7), (1, 6), (2, 5), (3, 4)],
    [(4, 10), (5, 3), (6, 2), (7, 1), (8, 9)],
    [(10, 9), (1, 8), (2, 7), (3, 6), (4, 5)],
    [(5, 10), (6, 4), (7, 3), (8, 2), (9, 1)],
]
PAIRINGS_8 = [
    [(1, 8), (2, 7), (3, 6), (4, 5)],
    [(8, 5), (6, 4), (7, 3), (1, 2)],
    [(2, 8), (3, 1), (4, 7), (5, 6)],
    [(8, 6), (7, 5), (1, 4), (2, 3)],
    [(3, 8), (4, 2), (5, 1), (6, 7)],
    [(8, 7), (1, 6), (2, 5), (3, 4)],
    [(4, 8), (5, 3), (6, 2), (7, 1)],
]
PAIRINGS_6 = [
    [(1, 6), (2, 5), (3, 4)],
    [(6, 4), (5, 3), (1, 2)],
    [(2, 6), (3, 1), (4, 5)],
    [(6, 5), (1, 4), (2, 3)],
    [(3, 6), (4, 2), (5, 1)],
]
PAIRINGS = {
    6: PAIRINGS_6,
    8: PAIRINGS_8,
    10: PAIRINGS_10,
}

ROUND_COUNT = 9
DIVISIONS = ['Staffel 1', 'Staffel 2', 'Staffel 3', 'Staffel 4']
PLAYERS = {
    'Staffel 1': [
        ('<lichess-name>', '<discord-name>'),
        ('<lichess-name>', '<discord-name>'),
        ('<lichess-name>', '<discord-name>'),
        ('<lichess-name>', '<discord-name>'),
        ('<lichess-name>', '<discord-name>'),
        ('<lichess-name>', '<discord-name>'),
        ('<lichess-name>', '<discord-name>'),
        ('<lichess-name>', '<discord-name>'),
        ('<lichess-name>', '<discord-name>'),
        ('<lichess-name>', '<discord-name>'),
    ],
    'Staffel 2': [
        ('<lichess-name>', '<discord-name>'),
        ('<lichess-name>', '<discord-name>'),
        ('<lichess-name>', '<discord-name>'),
        ('<lichess-name>', '<discord-name>'),
        ('<lichess-name>', '<discord-name>'),
        ('<lichess-name>', '<discord-name>'),
        ('<lichess-name>', '<discord-name>'),
        ('<lichess-name>', '<discord-name>'),
        ('<lichess-name>', '<discord-name>'),
        ('<lichess-name>', '<discord-name>'),
    ],
    'Staffel 3': [
        ('<lichess-name>', '<discord-name>'),
        ('<lichess-name>', '<discord-name>'),
        ('<lichess-name>', '<discord-name>'),
        ('<lichess-name>', '<discord-name>'),
        ('<lichess-name>', '<discord-name>'),
        ('<lichess-name>', '<discord-name>'),
        ('<lichess-name>', '<discord-name>'),
        ('<lichess-name>', '<discord-name>'),
        ('<lichess-name>', '<discord-name>'),
        ('<lichess-name>', '<discord-name>'),
    ],
    'Staffel 4': [
        ('<lichess-name>', '<discord-name>'),
        ('<lichess-name>', '<discord-name>'),
        ('<lichess-name>', '<discord-name>'),
        ('<lichess-name>', '<discord-name>'),
        ('<lichess-name>', '<discord-name>'),
        ('<lichess-name>', '<discord-name>'),
    ],
}

POTENTIAL_RESULTS = [
    ('Ergebnis', 'Punkte weiß', 'Punkte schwarz'),
    ('', '', ''),
    ('1-0', 1, 0),
    ('0-1', 0, 1),
    ('½-½', 0.5, 0.5),
    ('1-0 o.K.', 1, 0),
    ('0-1 o.K.', 0, 1),
    ('0-0 o.K.', 0, 0),
]


def _parse_date(s):
    return datetime.datetime.strptime(s, '%Y-%m-%d')


KW_START = 32
REGISTER_BY_STARTDATE = _parse_date('2022-08-09')
PLAY_UNTIL_DATE = _parse_date('2022-08-14')
SPECIAL_REGISTER_BY = {
    1: _parse_date('2022-08-15'),
    2: _parse_date('2022-08-16'),
}
SPECIAL_PLAY_UNTIL = {
    1: _parse_date('2022-08-21'),
}
DATE_FORMAT = 'd. mmmm yyyy'


def main():
    workbook = openpyxl.Workbook()

    # potential results
    potential_results = workbook.active
    potential_results.title = 'Mögliche Ergebnisse'
    for row_idx, row_data in enumerate(POTENTIAL_RESULTS, start=1):
        for col_idx, value in enumerate(row_data, start=1):
            potential_results.cell(column=col_idx, row=row_idx, value=value)
    potential_results.sheet_state = 'hidden'

    # Dates
    date_sheet = workbook.create_sheet('Daten')
    date_sheet.sheet_state = 'hidden'
    date_sheet['A1'] = 'Runde'
    date_sheet['B1'] = 'Kalenderwoche'
    date_sheet['C1'] = 'Termine vorgeschlagen bis: '
    date_sheet.column_dimensions['C'].width = 20
    date_sheet['D1'] = 'Letzter möglicher Spieltag: '
    register_by_date = REGISTER_BY_STARTDATE
    date_sheet.column_dimensions['D'].width = 25
    play_until_date = PLAY_UNTIL_DATE
    for round_num in range(1, ROUND_COUNT + 1):
        round_register_by = SPECIAL_REGISTER_BY.get(round_num) or register_by_date
        round_play_until = SPECIAL_PLAY_UNTIL.get(round_num) or play_until_date

        round_row = 1 + round_num
        date_sheet.cell(column=1, row=round_row, value=f'Runde {round_num}')
        date_sheet.cell(column=2, row=round_row, value=f'KW {KW_START + round_num}')
        c = date_sheet.cell(column=3, row=round_row)
        c.value = round_register_by
        c.number_format = DATE_FORMAT

        c = date_sheet.cell(column=4, row=round_row)
        c.value = round_play_until
        c.number_format = DATE_FORMAT

        register_by_date += datetime.timedelta(days=7)
        play_until_date += datetime.timedelta(days=7)

    # Main division sheets
    for division in DIVISIONS:
        ws = workbook.create_sheet(division)
        result_validation = DataValidation(
            type='list', formula1=f'{quote_sheetname(potential_results.title)}!$A$2:$A${len(POTENTIAL_RESULTS)+1}', allow_blank=True)
        ws.add_data_validation(result_validation)
        division_players = PLAYERS[division]
        division_pairings = PAIRINGS[len(division_players)]

        ws['A1'] = 'lichess'
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
        ws['C1'] = 'discord'
        ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=4)
        ws['E1'] = 'Rating'
        ws['F1'] = 'Punkte'

        MAIN_START = 13
        ws.cell(column=3, row=MAIN_START, value='Termin bis')
        ws.column_dimensions['C'].width = 18
        ws.cell(column=4, row=MAIN_START, value='gespielt bis')
        ws.column_dimensions['D'].width = 18
        announcement_column = 7 + len(division_pairings[0]) * 6 + 1
        ws.column_dimensions[get_column_letter(announcement_column)].width = 40
        point_cells = collections.defaultdict(list)  # player number (1-based) -> list of cells
        for round_idx, round_pairings in enumerate(division_pairings):
            round_player_cells = []
            round_row = MAIN_START + 1 + round_idx

            ws.row_dimensions[round_row].height = 15

            ws.cell(column=1, row=round_row, value=f"={quote_sheetname('Daten')}!A{round_idx + 2}")
            ws.cell(column=2, row=round_row, value=f"={quote_sheetname('Daten')}!B{round_idx + 2}")
            c = ws.cell(column=3, row=round_row, value=f"={quote_sheetname('Daten')}!C{round_idx + 2}")
            c.number_format = DATE_FORMAT
            c = ws.cell(column=4, row=round_row, value=f"={quote_sheetname('Daten')}!D{round_idx + 2}")
            c.number_format = DATE_FORMAT

            for pairing_idx, (white_num, black_num) in enumerate(round_pairings):
                pairing_start_column = 7 + pairing_idx * 6

                if round_idx == 0:
                    white_cell = ws.cell(column=pairing_start_column, row=MAIN_START, value='weiß')
                    white_cell.alignment = Alignment(horizontal='right')
                    ws.column_dimensions[get_column_letter(pairing_start_column)].width = 18
                    black_cell = ws.cell(column=pairing_start_column + 4, row=MAIN_START, value='schwarz')
                    black_cell.alignment = Alignment(horizontal='left')
                    ws.column_dimensions[get_column_letter(pairing_start_column + 4)].width = 18

                    result_column_letter = get_column_letter(pairing_start_column + 2)
                    result_validation.add(f'{result_column_letter}{round_row}:{result_column_letter}{round_row + len(division_pairings) - 1}')

                    ws.column_dimensions[get_column_letter(pairing_start_column + 1)].hidden = True
                    ws.column_dimensions[get_column_letter(pairing_start_column + 3)].hidden = True

                    ws.cell(column=announcement_column, row=MAIN_START, value='Ankündigung')

                white_cell = ws.cell(column=pairing_start_column, row=round_row, value=f'=A{white_num + 1}')
                white_cell.alignment = Alignment(horizontal='right')

                point_cells[white_num].append(f'{get_column_letter(pairing_start_column + 1)}{round_row}')
                point_cells[black_num].append(f'{get_column_letter(pairing_start_column + 3)}{round_row}')
                ws.cell(
                    column=pairing_start_column + 1, row=round_row,
                    value=(
                        f'=IF(ISBLANK({get_column_letter(pairing_start_column + 2)}{round_row});0;'
                        f'INDEX({quote_sheetname(potential_results.title)}!$A$2:$C${1 + len(POTENTIAL_RESULTS)},'
                        f' MATCH({get_column_letter(pairing_start_column + 2)}{round_row},'
                        f' {quote_sheetname(potential_results.title)}!$A$2:$A${1 + len(POTENTIAL_RESULTS)}, 0), 2))'))
                result_cell = ws.cell(column=pairing_start_column + 2, row=round_row, value='')
                result_cell.alignment = Alignment(horizontal='center')
                ws.cell(
                    column=pairing_start_column + 3, row=round_row,
                    value=(
                        f'=IF(ISBLANK({get_column_letter(pairing_start_column + 2)}{round_row});0;'
                        f'INDEX({quote_sheetname(potential_results.title)}!$A$2:$C${1 + len(POTENTIAL_RESULTS)},'
                        f' MATCH({get_column_letter(pairing_start_column + 2)}{round_row},'
                        f' {quote_sheetname(potential_results.title)}!$A$2:$A${1 + len(POTENTIAL_RESULTS)}, 0), 3))'))

                black_cell = ws.cell(column=pairing_start_column + 4, row=round_row, value=f'=A{black_num + 1}')
                black_cell.alignment = Alignment(horizontal='left')

                round_player_cells.append((
                    f'{get_column_letter(pairing_start_column)}{round_row}',
                    f'{get_column_letter(pairing_start_column + 4)}{round_row}'))

            # announcement string
            announcement_formula = (
                (f'="**Schachmattliga {division} " & A{round_row} & "**" & CHAR(10)'
                 f' & {quote_sheetname(date_sheet.title)}!C1 & TEXT(C{round_row}, "{DATE_FORMAT}") & CHAR(10)'
                 f' & {quote_sheetname(date_sheet.title)}!D1 & TEXT(D{round_row}, "{DATE_FORMAT}") & CHAR(10) & CHAR(10) & ') +
                (' & CHAR(10) & '.join(
                    f'{white_cell}'
                    f' & " @" & INDEX($A$2:$C${1 + len(division_players)}, MATCH({white_cell}, $A2:$A${1 + len(division_players)}, 0), 3)'
                    ' & " (weiß) - "'
                    f' & {black_cell} & " @"'
                    f' & INDEX($A$2:$C${1 + len(division_players)}, MATCH({black_cell}, $A2:$A${1 + len(division_players)}, 0), 3)'
                    f' & " (schwarz)"'
                    for white_cell, black_cell in round_player_cells
                ))
            )
            ws.cell(column=announcement_column, row=round_row, value=announcement_formula)

        for player_idx, player in enumerate(division_players):
            lichess_name, discord_name = player
            player_row = 2 + player_idx

            # lichess name
            lichess_cell = ws.cell(column=1, row=player_row, value=lichess_name)
            ws.merge_cells(start_row=player_row, start_column=1, end_row=player_row, end_column=2)

            # discord name
            ws.cell(column=3, row=player_row, value=discord_name)
            ws.merge_cells(start_row=player_row, start_column=3, end_row=player_row, end_column=4)

            # rating
            c = ws.cell(column=5, row=player_row)
            c.value = f'=IMPORTXML("https://phihag.de/2022/lichess-rating.php?user=" & {lichess_cell.coordinate}, "/rating")'

            # total points
            c = ws.cell(column=6, row=player_row, value='=' + '+'.join(point_cells[player_idx + 1]))


    vacation_sheet = workbook.create_sheet('Urlaub')
    vacation_sheet['A1'] = 'lichess-Name'
    vacation_sheet.column_dimensions['A'].width = 10
    vacation_sheet.column_dimensions['B'].width = 10
    vacation_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    vacation_sheet['C1'] = 'discord-Name'
    vacation_sheet.column_dimensions['C'].width = 10
    vacation_sheet.column_dimensions['D'].width = 10
    vacation_sheet.merge_cells(start_row=1, start_column=3, end_row=1, end_column=4)
    vacation_sheet['E1'] = 'Wunsch'
    vacation_sheet.column_dimensions['E'].width = 30

    # save workbook
    workbook.save('schachmattliga.xlsx')


if __name__ == '__main__':
    main()
